from __future__ import annotations

from dataclasses import replace
from pathlib import Path
import tempfile
import unittest

from openpyxl import load_workbook

from sand_bulletin.analytics import AnalyticsBlockedError, build_report_ready_metrics
from sand_bulletin.data_quality import validate_upload_batch
from sand_bulletin.dashboard.data import issue_frame, metric_frame
from sand_bulletin.ingestion import build_upload_batch_manifest
from sand_bulletin.ingestion.datasets import DatasetKind
from sand_bulletin.reports import generate_reports
from sand_bulletin.reports.html import render_bulletin_html
from sand_bulletin.reports.workbook import write_excel_workbook


DATA_DIR = Path(__file__).resolve().parents[1] / "docs" / "data"


class ReportGenerationTests(unittest.TestCase):
    def test_report_generation_blocks_raw_mode_on_high_severity_by_default(self) -> None:
        manifest = build_upload_batch_manifest(DATA_DIR, tenant_id="sand", country_code="RWA")

        with tempfile.TemporaryDirectory() as tmp:
            with self.assertRaises(AnalyticsBlockedError):
                generate_reports(manifest, Path(tmp), analysis_mode="raw")

    def test_excel_workbook_contains_expected_sheets(self) -> None:
        manifest = _clean_manifest()
        metrics = build_report_ready_metrics(manifest, validate_upload_batch(manifest))
        metrics_df = metric_frame(metrics)
        issues_df = issue_frame(validate_upload_batch(manifest))

        with tempfile.TemporaryDirectory() as tmp:
            path = Path(tmp) / "bulletin.xlsx"
            write_excel_workbook(path, metrics_df, issues_df, manifest)
            workbook = load_workbook(path)

        self.assertIn("summary_kpis", workbook.sheetnames)
        self.assertIn("facility_rankings", workbook.sheetnames)
        self.assertIn("district_trends", workbook.sheetnames)
        self.assertIn("nowcasts", workbook.sheetnames)
        self.assertIn("clinical_resolution", workbook.sheetnames)
        self.assertIn("data_quality_issues", workbook.sheetnames)
        self.assertIn("metric_trace", workbook.sheetnames)

    def test_html_bulletin_renders_core_sections(self) -> None:
        manifest = _clean_manifest()
        summary = validate_upload_batch(manifest)
        metrics = build_report_ready_metrics(manifest, summary)
        html = render_bulletin_html(metric_frame(metrics), issue_frame(summary), manifest, "test-run")

        self.assertIn("Quarterly Health Bulletin", html)
        self.assertIn("Executive Summary", html)
        self.assertIn("Data Quality Assessment", html)
        self.assertIn("Provincial Analysis", html)
        self.assertIn("Methodology and Limitations", html)
        self.assertIn("data:image/png;base64", html)

    def test_generate_reports_writes_html_and_excel(self) -> None:
        manifest = _clean_manifest()

        with tempfile.TemporaryDirectory() as tmp:
            generated = generate_reports(manifest, Path(tmp))

            self.assertTrue(generated.html_path.exists())
            self.assertTrue(generated.excel_path.exists())
            self.assertGreater(generated.metrics_count, 100)
            self.assertGreaterEqual(generated.issues_count, 0)
            self.assertIn(
                generated.pdf_status,
                {
                    "generated",
                    "skipped: WeasyPrint is not installed",
                    "skipped: WeasyPrint system dependencies are not available",
                },
            )

    def test_generate_reports_can_explicitly_override_sample_dq(self) -> None:
        manifest = build_upload_batch_manifest(DATA_DIR, tenant_id="sand", country_code="RWA")

        with tempfile.TemporaryDirectory() as tmp:
            generated = generate_reports(manifest, Path(tmp), allow_high_severity=True)

            self.assertTrue(generated.html_path.exists())
            self.assertTrue(generated.excel_path.exists())
            self.assertGreater(generated.issues_count, 0)


def _clean_manifest():
    manifest = build_upload_batch_manifest(DATA_DIR, tenant_id="sand", country_code="RWA")
    clinical_dataset = manifest.datasets[DatasetKind.CLINICAL_NEONATAL]
    clean_clinical = clinical_dataset.frame.drop_duplicates(
        subset=["facility_id", "reporting_month"],
        keep="first",
    ).reset_index(drop=True)
    manifest.datasets[DatasetKind.CLINICAL_NEONATAL] = replace(
        clinical_dataset,
        frame=clean_clinical,
    )

    governance_dataset = manifest.datasets[DatasetKind.GOVERNANCE]
    clean_governance = governance_dataset.frame.copy()
    clean_governance["hmis_reporting_completeness"] = clean_governance[
        "hmis_reporting_completeness"
    ].clip(lower=0.80)
    manifest.datasets[DatasetKind.GOVERNANCE] = replace(
        governance_dataset,
        frame=clean_governance,
    )
    return manifest


if __name__ == "__main__":
    unittest.main()
